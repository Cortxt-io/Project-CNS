"""Export node data to Excel (xlsx) format."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scripts.md_parser import read_all_nodes

EXPORTS_DIR = Path(__file__).resolve().parent.parent / "exports"

COLUMNS = [
    "Node",
    "Status",
    "MVP Stage",
    "Problem",
    "Solution",
    "Primary Audience",
    "Secondary Audience",
    "Assumptions",
    "Why Buy Not Build",
    "Cost SEK",
    "Value SEK",
    "ROI %",
    "Top Risk",
    "Updated",
]


def _first_n_chars(text: str, n: int = 100) -> str:
    text = text.strip()
    if len(text) <= n:
        return text
    return text[:n] + "..."


def _extract_audience(section_text: str, label: str) -> str:
    """Extract a labelled audience from the Target Audience section."""
    for line in section_text.splitlines():
        if label in line:
            return line.split(":", 1)[-1].strip().strip("*")
    return ""


def _join_list_items(section_text: str) -> str:
    """Join bullet-list items into a semicolon-separated string."""
    items = []
    for line in section_text.splitlines():
        line = line.strip()
        if line.startswith("- "):
            items.append(line[2:].strip())
    return "; ".join(items) if items else section_text.strip()


def _top_risk(section_text: str) -> str:
    """Return the first risk line from Risk Assessment."""
    for line in section_text.splitlines():
        line = line.strip()
        if line.startswith("- "):
            return line[2:].strip()
    return ""


def export_xlsx() -> Path:
    """Generate exports/MVP_comparison.xlsx from all node files.

    Returns the path to the generated file.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = EXPORTS_DIR / "MVP_comparison.xlsx"

    nodes = read_all_nodes()

    wb = Workbook()
    ws = wb.active
    ws.title = "MVP Comparison"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (meta, sections) in enumerate(nodes, 2):
        audience_text = sections.get("Target Audience", "")
        ws.cell(row=row_idx, column=1, value=meta.get("title", ""))
        ws.cell(row=row_idx, column=2, value=meta.get("status", ""))
        ws.cell(row=row_idx, column=3, value=meta.get("mvp_stage", ""))
        ws.cell(row=row_idx, column=4, value=_first_n_chars(sections.get("Problem", "")))
        ws.cell(row=row_idx, column=5, value=sections.get("Solution", "").strip())
        ws.cell(row=row_idx, column=6, value=_extract_audience(audience_text, "Primary"))
        ws.cell(row=row_idx, column=7, value=_extract_audience(audience_text, "Secondary"))
        ws.cell(row=row_idx, column=8, value=_join_list_items(sections.get("Assumptions to Validate", "")))
        ws.cell(row=row_idx, column=9, value=_join_list_items(sections.get("Why Buy Instead of Build?", "")))
        kind = meta.get("kind")
        if kind is None:
            ws.cell(row=row_idx, column=10, value=meta.get("cost_sek", 0))
            ws.cell(row=row_idx, column=11, value=meta.get("value_sek", 0))
            ws.cell(row=row_idx, column=12, value=meta.get("roi_percent", 0))
        else:
            ws.cell(row=row_idx, column=10, value=meta.get("stage", ""))
            ws.cell(row=row_idx, column=11, value=meta.get("kind", ""))
            ws.cell(row=row_idx, column=12, value="")  # No ROI for kind-aware
        ws.cell(row=row_idx, column=13, value=_top_risk(sections.get("Risk Assessment", "")))
        ws.cell(row=row_idx, column=14, value=str(meta.get("updated", "")))

    # Auto-fit column widths (approximation)
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 4, 15)

    wb.save(str(output_path))
    return output_path
